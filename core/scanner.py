"""xlsx/xls文件扫描器 - 递归扫描目录并提取子表名称"""
import os
import time
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
import xlrd

# XML 命名空间
NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

class XlsxScanner:
    def __init__(self, max_workers: int = 8):
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
        self.max_workers = max_workers

    def is_xlsx_file(self, filepath: str) -> bool:
        """检查是否为支持的表格文件"""
        ext = os.path.splitext(filepath)[1].lower()
        return ext in self.supported_extensions

    @staticmethod
    def _is_xls_format(filepath: str) -> bool:
        """判断是否为旧版 .xls 格式"""
        return os.path.splitext(filepath)[1].lower() == '.xls'

    def get_sheet_names(self, filepath: str) -> List[str]:
        """获取表格文件的所有子表名称"""
        if self._is_xls_format(filepath):
            return self._get_sheet_names_xls(filepath)

        try:
            # xlsx/xlsm 是 zip 文件，直接读取 workbook.xml 获取 sheet 名称
            with zipfile.ZipFile(filepath, 'r') as zf:
                with zf.open('xl/workbook.xml') as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    sheets = root.findall('.//main:sheet', NS)
                    if sheets:
                        return [s.get('name', f'Sheet{i+1}') for i, s in enumerate(sheets)]
            return []
        except Exception as e:
            # 备用方案：使用 openpyxl
            return self._get_sheet_names_slow(filepath)

    def _get_sheet_names_xls(self, filepath: str) -> List[str]:
        """使用 xlrd 获取 .xls 文件的子表名称"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            return wb.sheet_names()
        except Exception as e:
            print(f"警告: 读取 .xls 文件失败 {filepath}: {e}")
            return []

    def _get_sheet_names_slow(self, filepath: str) -> List[str]:
        """备用方式获取子表名称（使用 openpyxl）"""
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"警告: 无法读取文件 {filepath}: {e}")
            return []

    def extract_cell_texts(self, filepath: str, sheet_names: List[str],
                           max_chars_per_sheet: int = 50000) -> List[str]:
        """
        提取每个 sheet 的单元格内容（拼接为字符串），用于索引。
        打开工作簿一次，遍历所有请求的 sheet。
        返回与 sheet_names 平行的字符串列表。
        """
        if self._is_xls_format(filepath):
            return self._extract_cell_texts_xls(filepath, sheet_names, max_chars_per_sheet)

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            results = []
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    results.append('')
                    continue
                ws = wb[sheet_name]
                parts = []
                char_count = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            val = str(cell.value)
                            parts.append(val)
                            char_count += len(val) + 1
                            if char_count > max_chars_per_sheet:
                                break
                    if char_count > max_chars_per_sheet:
                        break
                results.append(' '.join(parts))
            wb.close()
            return results
        except Exception as e:
            print(f"警告: 提取单元格文本失败 {filepath}: {e}")
            return [''] * len(sheet_names)

    def _extract_cell_texts_xls(self, filepath: str, sheet_names: List[str],
                                 max_chars_per_sheet: int = 50000) -> List[str]:
        """xlrd 版本：提取 .xls 文件的单元格内容"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            results = []
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheet_names():
                    results.append('')
                    continue
                ws = wb.sheet_by_name(sheet_name)
                parts = []
                char_count = 0
                for row_idx in range(ws.nrows):
                    for col_idx in range(ws.ncols):
                        val = ws.cell_value(row_idx, col_idx)
                        if val is not None and val != '':
                            s = str(val)
                            parts.append(s)
                            char_count += len(s) + 1
                            if char_count > max_chars_per_sheet:
                                break
                    if char_count > max_chars_per_sheet:
                        break
                results.append(' '.join(parts))
            wb.release_resources()
            return results
        except Exception as e:
            print(f"警告: 提取 .xls 单元格文本失败 {filepath}: {e}")
            return [''] * len(sheet_names)

    def read_sheet_preview(self, filepath: str, sheet_name: str,
                           max_rows: int = 20, max_cols: int = 50) -> List[List[str]]:
        """
        读取单个 sheet 的前 max_rows 行 × max_cols 列数据。
        用于预览面板展示。
        """
        if self._is_xls_format(filepath):
            return self._read_sheet_preview_xls(filepath, sheet_name, max_rows, max_cols)

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[sheet_name]
            data = []
            row_count = 0
            for row in ws.iter_rows(max_col=max_cols, max_row=max_rows):
                data.append([str(cell.value) if cell.value is not None else '' for cell in row])
                row_count += 1
                if row_count >= max_rows:
                    break
            wb.close()
            return data
        except Exception as e:
            print(f"警告: 读取预览数据失败 {filepath}: {e}")
            return []

    def _read_sheet_preview_xls(self, filepath: str, sheet_name: str,
                                 max_rows: int = 20, max_cols: int = 50) -> List[List[str]]:
        """xlrd 版本：读取 .xls 文件的预览数据"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            if sheet_name not in wb.sheet_names():
                wb.release_resources()
                return []
            ws = wb.sheet_by_name(sheet_name)
            data = []
            for row_idx in range(min(ws.nrows, max_rows)):
                row_data = []
                for col_idx in range(min(ws.ncols, max_cols)):
                    val = ws.cell_value(row_idx, col_idx)
                    row_data.append(str(val) if val is not None and val != '' else '')
                data.append(row_data)
            wb.release_resources()
            return data
        except Exception as e:
            print(f"警告: 读取 .xls 预览数据失败 {filepath}: {e}")
            return []

    def scan_file(self, filepath: str) -> Tuple[str, float, List[str]] | None:
        """扫描单个文件，返回(文件名, 修改时间, 子表列表)"""
        if not self.is_xlsx_file(filepath):
            return None

        try:
            stat = os.stat(filepath)
            modified_time = stat.st_mtime
            filename = os.path.basename(filepath)
            sheet_names = self.get_sheet_names(filepath)
            return (filename, modified_time, sheet_names)
        except Exception as e:
            print(f"警告: 扫描文件失败 {filepath}: {e}")
            return None

    def scan_directory(self, directory: str, progress_callback: Callable = None) -> List[Tuple[str, str, float, List[str]]]:
        """
        递归扫描目录下的所有xlsx文件
        返回: [(filename, filepath, modified_time, sheet_names), ...]
        """
        results = []

        for root, dirs, files in os.walk(directory):
            # 跳过隐藏目录
            dirs[:] = [d for d in dirs if not d.startswith('.')]

            for filename in files:
                if not self.is_xlsx_file(filename):
                    continue

                filepath = os.path.join(root, filename)
                result = self.scan_file(filepath)

                if result:
                    results.append((result[0], filepath, result[1], result[2]))

                if progress_callback:
                    progress_callback(len(results))

        return results

    def scan_directory_incremental(self, directory: str, index_manager, progress_callback: Callable = None) -> Tuple[int, int, int]:
        """
        增量扫描目录，只更新有变化的文件（多线程并发）
        返回: (新增文件数, 更新文件数, 删除文件数)
        """
        # 获取所有需要扫描的文件
        all_files = {}
        for root, dirs, files in os.walk(directory):
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for filename in files:
                if self.is_xlsx_file(filename):
                    filepath = os.path.join(root, filename)
                    all_files[filepath] = True

        # 获取索引中已存在的文件
        indexed_files = {}
        for file_info in index_manager.get_all_files():
            indexed_files[file_info['filepath']] = file_info['modified_time']

        added = 0
        updated = 0
        deleted = 0

        # 需要处理的文件列表
        files_to_process = []
        for filepath in all_files:
            try:
                stat = os.stat(filepath)
                modified_time = stat.st_mtime
                filename = os.path.basename(filepath)
                existing = index_manager.get_file_info(filepath)

                if existing is None:
                    files_to_process.append((filepath, filename, modified_time, 'add'))
                elif existing[1] != modified_time:
                    files_to_process.append((filepath, filename, modified_time, 'update'))
            except Exception as e:
                print(f"警告: 获取文件信息失败 {filepath}: {e}")

        # 收集需要写入索引的数据
        pending_updates = []  # [(filepath, filename, modified_time, sheet_names), ...]

        # 并发处理文件
        total_files = len(files_to_process)
        if progress_callback:
            progress_callback(0, total_files)

        if files_to_process:
            processed = 0
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {
                    executor.submit(self.get_sheet_names, filepath): (filepath, filename, modified_time)
                    for filepath, filename, modified_time, op_type in files_to_process
                }

                for future in as_completed(futures):
                    filepath, filename, modified_time = futures[future]
                    processed += 1
                    try:
                        sheet_names = future.result()
                        if sheet_names:
                            pending_updates.append((filename, filepath, modified_time, sheet_names))
                    except Exception as e:
                        print(f"警告: 处理文件失败 {filepath}: {e}")
                    finally:
                        if progress_callback:
                            progress_callback(processed, total_files)

        # 主线程写入数据库（线程安全）
        for filename, filepath, modified_time, sheet_names in pending_updates:
            index_manager.add_file(filename, filepath, modified_time, sheet_names)
            existing = index_manager.get_file_info(filepath)
            if existing:
                updated += 1
            else:
                added += 1

        # 处理已删除的文件
        for filepath in indexed_files:
            if filepath not in all_files:
                index_manager.delete_file(filepath)
                deleted += 1

        return (added, updated, deleted)